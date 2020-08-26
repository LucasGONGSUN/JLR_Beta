import json
import openpyxl
import pymongo
import smtplib
import schedule
import time
from datetime import datetime
from math import ceil
from pymongo import errors
from email.mime.text import MIMEText
from email.header import Header



# Settings
# --- Files --- #
FileJsonName_date = 'Calendar.json'
FileExcelName_date = 'Calendar.xlsx'

FileJsonName_content = 'JLR_Contents.json'
FileExcelName_content = 'JLR_Contents.xlsx'

FileJsonName_user = 'UserInfo.json'
FileExcelName_user = 'Users.xlsx'

# --- MongoDB --- #
DatabaseName = 'JLR_Beta'

host_u = 'mongodb://localhost:27017'
#host_u = "mongodb+srv://lucassl3:lucas112358@alphajune-czcux.azure.mongodb.net/<dbname>" \
#         "?retryWrites=true&w=majority"
colname_u = 'JLR_User'

host_c = 'mongodb://localhost:27017'
#host_c = "mongodb+srv://lucassl3:lucas112358@alphajune-czcux.azure.mongodb.net/" \
#         "<dbname>?retryWrites=true&w=majority"
colname_c = 'JLR_Local'

# --- Functions --- #
KanzisNum = 15
TodayDate = datetime.today()
Today = datetime(TodayDate.year, TodayDate.month, TodayDate.day, 0,0,0)


# ##### Define class Learner() ##### #
class Learner:

    def GetStartDay(self, UserList, LearnerNumber):
        """
            Get StartDay from UserList
        """
        StartDay = UserList[LearnerNumber]['FirstDay']
        return StartDay

    def GetDayNumber(self, StartDay, Today):
        """
            Get DayNumber of learning contents of today
        """
        DayFrom = datetime.strptime(StartDay, '%Y-%m-%d %H:%M:%S')
        DayNumber = int((Today - DayFrom).days) + 1
        return DayNumber

    def GetTodayList(self, CheckList, DayNumber):
        """
            Get today's list numbers of contents
        """
        TodayList = CheckList[DayNumber]
        return TodayList

    def GetTodayContent(self, ContentList, TodayList):
        """
            Get learning contents for today from ContentList
        """
        # Get New contents for today
        TodayNew = {}
        for i in TodayList[:1]:
            try:
                TodayNew[i] = ContentList[i]
            except KeyError:
                TodayNew['无'] = '无新学习内容'
                continue

        # Get Review contents for today
        TodayReview = {}
        for i in TodayList[1:]:
            try:
                TodayReview[i] = ContentList[i]
            except KeyError:
                continue
        return TodayNew, TodayReview

    def GetMailInfo(self, UserList, LearnerNumber):
        """
            Get MailAddr from UserList
        """
        MailTo = str(UserList[LearnerNumber]['ID'])
        MailAddr = UserList[LearnerNumber]['MailAddr']
        return MailTo, MailAddr

    def CreateMailContent(self, MailTo, TodayNew, TodayReview):
        """
            Create mail content for pushing
        """
        # Get numbers and contents for today's learning
        NewContent = []
        NewContent.append("\n\n\n>>> Today's Learning Contents <<<\n")
        for nums, contents in TodayNew.items():
            entry_num = "\n\n=== Learning Contents %s ===\n" % nums
            NewContent.append(entry_num)

            # Get Kanzi and DETAILS_DICT from CONTENTS_DICT
            try:
                for kanzis, details in contents.items():
                    entry_kanzi = '\n --- %s --- \n' % kanzis
                    NewContent.append(entry_kanzi)

                    # Get details from DETAILS_DICT
                    for keys, values in details.items():
                        entry_keys_values = '%s : %s \n' % (keys, values)
                        NewContent.append(entry_keys_values)
            except AttributeError:
                NewContent.append('\nNo new learnning contents for today.')

        # Get numbers and contents for today's reviewing
        ReviewContent = []
        ReviewContent.append("\n\n\n>>> Today's Reviewing Contents <<<\n")
        for nums, contents in TodayReview.items():
            entry_num = "\n\n=== Reviewing Contents %s ===\n" % nums
            ReviewContent.append(entry_num)

            # Get Kanzi and DETAILS_DICT from CONTENTS_DICT
            for kanzis, details in contents.items():
                entry_kanzi = '\n --- %s --- \n' % kanzis
                ReviewContent.append(entry_kanzi)

                # Get details from DETAILS_DICT
                for keys, values in details.items():
                    entry_keys_values = '%s : %s \n' % (keys, values)
                    ReviewContent.append(entry_keys_values)

        # Combine together
        LearningContent_str = " ".join(NewContent) + " ".join(ReviewContent)
        Greeting = '\nHello %s :' % MailTo
        MailContent = Greeting + LearningContent_str
        return MailContent


# ###### Load learning schedule from Excel to Json ###### #
# --- 1 Save calendar to Json --- #
def CreateSchedule():
    global CheckList

    # Settings
    WB = openpyxl.load_workbook('Calendar.xlsx')
    WS = WB['Jap']
    rows = WS.max_row
    cols = WS.max_column
    NumberList = []
    NoteList = []
    CheckList = {}

    # Read No. from calendar
    for row in WS.iter_rows(min_row=2, max_row=rows, max_col=1):
        for cell in row:
            NumberList.append(cell.value)

    # Read notelist from calendar
    for row in WS.iter_rows(min_row=2, max_row=rows, min_col=2, max_col=cols):
        new_row = []
        for cell in row:
            new_row.append(cell.value)
        NoteList.append(new_row)

    # Combine Number and notelist
    for i in range(184):
        CheckList[str(NumberList[i])] = NoteList[i]

    # Save to Json file
    with open(FileJsonName_date, 'w') as dw:
        json.dump(CheckList, dw, indent=2)
    print('\n>>> Calendar is saved to Json file %s. <<<' %FileJsonName_date)


# ###### User Management ###### #
# --- 2.1 Show user list --- #
def ShowUserList():
    global UserList

    # Show List of User Info
    print('\n===== User Info =====')
    try:
        for nums, infos in UserList.items():
            print('\n --- %s ---' % str(nums))
            for keys, values in infos.items():
                print('%s : %s' % (keys, values))
    except NameError:
        print('\nNo records, pleas load from Json file.')


# --- 2.2 Add a New user --- #
def AddNewUser():
    global UserList

    # Show List of User Info
    print('\n=== The followings are users already in the list. ===\n')
    for nums, infos in UserList.items():
        print('%s : %s' % (nums, infos['ID']))
    TotalNumber = int(len(UserList))
    NextNumber = str(TotalNumber + 1)
    print('\nTotal numbers: %s' % str(TotalNumber))

    # Input new user
    AddCheckPoint = input('\nGoing to input User No.%s , enter Y to continue: ' % NextNumber)
    while AddCheckPoint == 'Y':
        # Input info
        LearnerNumber = 'JLR-' + NextNumber
        ID = input('ID: ')
        FirstDay_M = int(input('Month of FirstDay: '))
        FirstDay_D = int(input('Day of FirstDay: '))
        MailAddr = input('MailAddr: ')

        # Save as entries
        NewEntry = {}
        NewEntry['LearnerNumber'] = LearnerNumber
        NewEntry['ID'] = ID
        NewEntry['FirstDay'] = str(datetime(2020, FirstDay_M, FirstDay_D, 0, 0, 0))
        NewEntry['MailAddr'] = MailAddr

        # Save check
        print('\nNew User Info:\n  LearnerNumber: %s\n  ID: %s\n  FirstDay: %s\n  MailAddr: %s'
              % (NewEntry['LearnerNumber'], NewEntry['ID'], NewEntry['FirstDay'], NewEntry['MailAddr']))

        SaveCheckPoint = input("\nPlease check user's information, enter Y to save or R to re-input: ")

        # Save to Json
        if SaveCheckPoint == 'Y':
            UserList[LearnerNumber] = NewEntry
            with open(FileJsonName_user, 'w') as uw:
                json.dump(UserList, uw, indent=2)
            print('\n>>> Learner %s: %s is saved to user list. <<<' % (LearnerNumber, ID))
            AddCheckPoint = 'N'

        # Input again
        elif SaveCheckPoint == 'R':
            print('\nGoing to input User No.%s again: ' % NextNumber)
            continue

        # Break
        else:
            print('\nMission About!')
            break


# --- 2.3 Import User List from Excel to Json --- #
def Excel2Json_U():
    global UserList

    # Settings
    wb = openpyxl.load_workbook(FileExcelName_user)
    ws = wb['JLR']
    rows = ws.max_row
    cols = ws.max_column
    Headers = []
    UserList = {}

    # Load Headers
    for row in ws.iter_rows(max_row=1, max_col=cols):
        for cell in row:
            Headers.append(cell.value)

    # Load details of Users
    for row in ws.iter_rows(min_row=2, max_row=rows, max_col=cols):
        LearnerNumber = row[1].value
        ID = row[2].value
        FirstDay = row[3].value
        MailAddr = row[4].value

        # Save as entries
        NewEntry = {}
        NewEntry[Headers[1]] = LearnerNumber
        NewEntry[Headers[2]] = ID
        NewEntry[Headers[3]] = str(FirstDay)
        NewEntry[Headers[4]] = MailAddr
        UserList[LearnerNumber] = NewEntry

    # Save to Json file
    with open(FileJsonName_user, 'w') as uw:
        json.dump(UserList, uw, indent=2)
    print('\nAll User Informations are saved to Json file %s.' %FileJsonName_user)


# --- 2.4 Export User List From Json to Excel --- #
def Json2Excel_U():
    global UserList

    # Settings
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'JLR'

    # Create list for all user info
    XlsWriterEntries = []  ### Create new list for all items
    IndexNumber = 1
    for users, details in UserList.items():  ### Get LearnerNumber and details
        XlsUserEntries = []  ### Create new list for a new user
        XlsUserEntries.append(IndexNumber)
        IndexNumber += 1
        for keys, values in details.items():  ### Get details
            XlsUserEntries.append(values)  ### Append values to UserEntry
        XlsWriterEntries.append(XlsUserEntries)  ### Append User to XlsWriterEntry

    # Save to Excel file
    headers = ['Index No.', 'LearnerNumber', 'ID', 'FirstDay', 'MailAddr']
    ws.append(headers)
    for i in XlsWriterEntries:
        ws.append(i)
    try:
        wb.save(FileExcelName_user)
        print('\n>>> User informations in Json are all saved to Excel %s. <<<' %FileExcelName_user)
        wb.close()
    except PermissionError:
        print('\nThe Excel file %s is running, please close it and try again.' %FileExcelName_user)


# --- 2.5 Upload User List to MongoDB --- #
def Json2MongoDB_U(HostAddr, CollectionName):
    global UserList

    # Get connection to MongoDB
    client = pymongo.MongoClient(HostAddr)
    print('\nMaking a connection with MongoClient ...')
    db = client[DatabaseName]
    print('Getting Database - %s ...' % DatabaseName)
    collection = db[CollectionName]
    print('Getting a Collection - %s ...\n' % CollectionName)

    # Create entry for posting
    for ln, contents in UserList.items():
        PostEntry = {}
        PostEntry['LearnerNumber'] = ln
        PostEntry['Contents'] = contents
        PostEntry['tags'] = 'JLR'
        PostEntry['date'] = datetime.today()

        # Post to MangoDB
        post_id = collection.insert_one(PostEntry).inserted_id
        print('Posted: %s - %s' % (PostEntry['LearnerNumber'], PostEntry['Contents']))


# --- 2.6 Download User List from MongoDB --- #
def MongoDB2Json_U(HostAddr, CollectionName):
    global UserList

    # Connect to MongoDB
    client = pymongo.MongoClient(HostAddr)
    print('\nMaking a connection with MongoClient ...')
    db = client[DatabaseName]
    print('Getting Database - %s ...' % DatabaseName)
    collection = db[CollectionName]
    print('Getting a Collection - %s\n' % CollectionName)

    # Download all contents
    AllUsers = collection.find()

    # Save Json contents
    UserList = {}
    for content in AllUsers:
        EntryNum = content['LearnerNumber']
        EntryContent = content['Contents']
        UserList[EntryNum] = EntryContent

    # Save to Json file
    with open(FileJsonName_user, 'w') as uw:
        json.dump(UserList, uw, indent=2)
    print('\n>>> All Users from MongoDB are saved to Json file. %s <<<' % FileJsonName_user)


# ###### Contents Management ###### #
# --- 3.1 Load and Inquire for Entries --- #
def Inquiry_C():
    global ContentList

    # Load Json file
    try:
        with open(FileJsonName_content, 'r') as cr:
            ContentList = json.load(cr)
    except FileNotFoundError:
        print('\n File Not Found. ')
        return

    # Show list of <Learning Contents No.>
    print("\n=== The followings are all the <Learning Contents No.> ===")
    NumsCount = 1
    for nums, dicts in ContentList.items():
        print(nums, end='  ')
        NumsCount += 1
        if NumsCount == 11:
            print('\n', end='')
            NumsCount = 1
    print('\n')

    # Input number of start
    InputCheckPoint_start = True
    while InputCheckPoint_start:
        try:
            ListNumberStart = int(input("Please enter the 'BEGINNING No.' (1-150): "))
            InputCheckPoint_start = False
        except ValueError:
            print('Please input an integrate number.')
            continue

    # Input number of end
    InputCheckPoint_end = True
    while InputCheckPoint_end:
        try:
            ListNumberEnd = int(input("Please enter the 'END No.' (1-150): "))
            InputCheckPoint_end = False
        except ValueError:
            print('Please input an integrate number.')
            continue

    # Print inquired contents
    try:
        for i in range(ListNumberStart, ListNumberEnd+1):
            print('\n\n===== Here are the contents of No.%s =====\n' %str(i))
            for keys, values in ContentList['No.'+str(i)].items():
                print(keys + ': ' + str(values))
    except KeyError:
        print('The List No. is out of range. Please try again.\n')


# --- 3.2 Input A New Entry --- #
def AddNewEntry(ContentList):
    global EntryNum, ContentCheckPoint

    # Get numbers
    LCNumberList = []
    print("\nThe following <Learning Contents No.> are already in the list: ")
    NumsCount = 1
    for nums, values in ContentList.items():
        print(nums, end=' ')
        NumsCount += 1
        if NumsCount == 16:
            print('\n')
        LCNumberList.append(nums)

    # Input new number and then CHECK
    ContentCheckPoint = 'Y'
    while ContentCheckPoint == 'Y':
        EntryNum = 'No.' + input('\n\nPlease input the number of NEW EntryList: ')
        if EntryNum in LCNumberList:
            print('Warning: This List Number is already in the list!')
            continue
        else:
            InputNewEntry(ContentList)


def InputNewEntry(ContentList):
    global EntryList, ContentCheckPoint
    EntryList = {}
    StartCheckPoint = input('Going to input EntryList %s, enter Y to continue: ' % EntryNum)
    if StartCheckPoint == 'Y':
        for i in range(1, 16):
            print(' >>> Now is inputting word No.%s <<< ' %i)

            #Input detailed contents
            kanzi = input('漢字：')
            kundoku = input('訓読：')
            kundokurei = input('訓読例：')
            onyomi = input('音読み：')
            onyomirei = input('音読み例：')

            #Save as entries
            NewEntry = {}
            NewEntry['漢字'] = kanzi
            NewEntry['訓読'] = kundoku
            NewEntry['訓読例'] = kundokurei
            NewEntry['音読み'] = onyomi
            NewEntry['音読み例'] = onyomirei
            EntryList[kanzi] = NewEntry

            #Check
            for keys, values in EntryList.items():
                print(keys + ': ' + str(values))
            ContinueCheckPoint = input('Above is word No.%s , enter N to abort: ' %i)
            if ContinueCheckPoint == 'N':
                print('\n>>> Mission Abort. See you next time. <<<')
                break
            else:
                continue

        # Combine EntryNum and EntryList, add to ContentList
        ContentList[EntryNum] = EntryList
        ContentCheckPoint = 'N'
    else:
        print('\n>>> Mission Quit. See you next time. <<<')
        ContentCheckPoint = 'N'

    with open(FileJsonName_content, 'w') as file_obj_w:
        json.dump(ContentList, file_obj_w, indent=2)
    print('\n>>> List %s is saved. Mission Completed! <<<' %EntryNum)


# --- 3.3 Continue An Unfinished Entry --- #
def ContinueEntry(ContentList):
    global EntryNum, ContentCheckPoint

    # Get numbers
    print('\n>>> Mission Start! <<<\n')
    LCNumberList = []
    print("\nThe following <Learning Contents No.> are already in the list: ")
    NumsCount = 1
    for nums, values in ContentList.items():
        print(nums, end=' ')
        NumsCount += 1
        if NumsCount == 16:
            print('\n')
        LCNumberList.append(nums)

    # Input a number and then CHECK
    ContentCheckPoint = 'Y'
    while ContentCheckPoint == 'Y':
        EntryNum = 'No.' + input('\n\nPlease input the number of list you want to CONTINUE: ')
        if EntryNum not in LCNumberList:
            print('Warning: This List Number is not in the list!')
            continue
        else:
            ContinueInput(ContentList)


def ContinueInput(ContentList):
    global EntryList, ContentCheckPoint

    # Get and print list contents
    EntryList = ContentList[EntryNum]
    print('\nHere are the contents of List %s :\n' %EntryNum)
    for keys, values in EntryList.items():
        print(keys + ': ' + str(values))

    # Continue to input
    StartCheckPoint = input('\nGoing to CONTINUE EntryList %s, enter N to abort: ' % EntryNum)
    if StartCheckPoint == 'N':
        print('\n>>> Mission Quit. See you next time. <<<')
        ContentCheckPoint = 'N'
    else:
        for i in range(len(ContentList[EntryNum])+1, 16):
            print(' >>> Now is inputting word No.%s <<< ' %i)

            #Input detailed contents
            kanzi = input('漢字：')
            kundoku = input('訓読：')
            kundokurei = input('訓読例：')
            onyomi = input('音読み：')
            onyomirei = input('音読み例：')

            #Save as entries
            NewEntry = {}
            NewEntry['漢字'] = kanzi
            NewEntry['訓読'] = kundoku
            NewEntry['訓読例'] = kundokurei
            NewEntry['音読み'] = onyomi
            NewEntry['音読み例'] = onyomirei
            EntryList[kanzi] = NewEntry

            #Check
            for keys, values in EntryList.items():
                print(keys + ': ' + str(values))
            ContinueCheckPoint = input('Above is word No.%s , enter N to abort: ' %i)
            if ContinueCheckPoint == 'N':
                print('\n>>> Mission Abort. See you next time. <<<')
                break
            else:
                continue

        # Combine EntryNum and EntryList, add to ContentList
        ContentList[EntryNum] = EntryList
        SaveJsonFile_C(ContentList)
        ContentCheckPoint = 'N'


# --- 3.4 Correct Values --- #
def SearchValue(ContentList):
    global ModifyTo, ListToModify, KanziToModify, ItemToModify

    # Get List
    print('\n>>> Mission Start! <<<\n')
    ListToModify = 'No.' + input('\nPlease input the list number: ')
    print('\n=== Here are the contents of List %s ===' % ListToModify)
    for keys, values in ContentList[ListToModify].items():
        print(keys + ': ' + str(values))

    # Get Kanzi
    KanziToModify = input('\nPlease input the Kanzi you want to modify: ')
    print('\n=== Here are the items of %s: ===' % KanziToModify)
    for keys, values in ContentList[ListToModify][KanziToModify].items():
        print(keys + ': ' + str(values))

    # Get Value
    ItemToModify = input('\nPlease input the Item you want to modify: ')
    OriginValue = ContentList[ListToModify][KanziToModify][ItemToModify]
    print("\nThe ENTRY you want to correct is %s:" % ContentList[ListToModify][KanziToModify])
    print("The VALUE you want to correct is %s:" % OriginValue)

    # Input Right Value
    ModifyCheckPoint = input('\nDo you want to correct the value? Enter Y to correct: ')
    if ModifyCheckPoint == 'Y':
        ModifyTo = input('Please input the right value: ')
        ModifyToCheck = input("\nDo you want to change '%s' to '%s' ? Enter Y to DO THAT: " % (OriginValue, ModifyTo))
        if ModifyToCheck == 'Y':
            CorrectValue(ContentList)
        else:
            print('\nMission Abort!')
    else:
        print('\nMission Quit!')


def CorrectValue(ContentList):
    ContentList[ListToModify][KanziToModify][ItemToModify] = ModifyTo
    print('\nThe VALUE is corrected to %s.' %ModifyTo)
    print("The ENTRY now is: %s" %ContentList[ListToModify][KanziToModify])
    SaveCheckPoint = input("\nEnter Y to save the Correction: ")
    if SaveCheckPoint == 'Y':
        SaveJsonFile_C(ContentList)
    else:
        print('\n>>> Mission Canceled <<<')


# --- 3.5 Import Contents From Excel to Json --- #
def Excel2Json_C():
    global ContentList

    # Settings
    wb = openpyxl.load_workbook(FileExcelName_content)
    ws = wb['JLContents']
    rows = ws.max_row
    cols = ws.max_column
    Headers = []
    EntryList = {}
    ContentList = {}

    # Load Headers
    for row in ws.iter_rows(max_row=1, max_col=cols):
        for cell in row:
            Headers.append(cell.value)

    # Load details of Kanzi
    for row in ws.iter_rows(min_row=1, max_row=rows, max_col=cols):
        kanzi = row[1].value
        kundoku = row[2].value
        kundokurei = row[3].value
        onyomi = row[4].value
        onyomirei = row[5].value

        # Save as entries
        NewEntry = {}
        NewEntry[Headers[1]] = kanzi
        NewEntry[Headers[2]] = kundoku
        NewEntry[Headers[3]] = kundokurei
        NewEntry[Headers[4]] = onyomi
        NewEntry[Headers[5]] = onyomirei
        EntryList[kanzi] = NewEntry

    # Convert dict contents to a list of key-value pairs
    KanziItems = list(EntryList.items())

    # Group entries by number of KanziNum
    for i in range(ceil(len(KanziItems) / KanzisNum)):
        EntryListNum = 'No.' + str(i + 1)
        GroupEntryList = {}
        for x in range(i * KanzisNum + 1, (i + 1) * KanzisNum + 1):
            try:
                GroupEntryList[KanziItems[x][0]] = KanziItems[x][1]
            except IndexError:
                break
        ContentList[EntryListNum] = GroupEntryList

    # Save to Json file
    SaveJsonFile_C(ContentList)


# --- 3.6 Export Contents From Json to Excel --- #
def Json2Excel_C(ContentList):
    # Settings
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'JLContents'
    headers = ['List No.', '漢字', '訓読', '訓読例', '音読み', '音読み例']
    ws.append(headers)

    # Create list for all kanzis and details
    XlsListEntries = []  ### Create new list for all items
    for nums, dicts in ContentList.items():  ### Get list number
        for kanzis, details in dicts.items():  ### Get kanzis and details
            XlsKanziEntries = []  ### Create new list for a new kanzi
            XlsKanziEntries.append(nums)  ### Append List No. to KanziEntry
            for keys, values in details.items():  ### Get values
                XlsKanziEntries.append(values)  ### Append values to KanziEntry
            XlsListEntries.append(XlsKanziEntries)  ### Append KanziEntry to XlsEntry

    # Save to Excel file
    for i in XlsListEntries:
        ws.append(i)
    try:
        wb.save(FileExcelName_content)
        print('\n>>> Contents in Json are all saved to Excel %s. <<<' % FileExcelName_content)
        wb.close()
    except PermissionError:
        print('\nThe Excel file %s is running, please close it and try again.' % FileExcelName_content)


# --- 3.7 Upload Contents to MangoDB --- #
def Json2MongoDB_C(HostAddr, CollectionName):
    global ContentList

    # Get connection with MongoDB
    client = pymongo.MongoClient(HostAddr)
    print('\nMaking a connection with MongoClient ...')
    db = client[DatabaseName]
    print('Getting Database - %s ...' % DatabaseName)
    collection = db[CollectionName]
    print('Getting a Collection - %s ...\n' % CollectionName)

    # Create entry for posting
    for nums, values in ContentList.items():
        for kanzi, contents in values.items():
            PostEntry = {}
            try:
                id = 'Kanzi' + str(ord(kanzi))
                PostEntry['_id'] = id
            except TypeError:
                print('Kanzi is null!')
                continue
            PostEntry['Kanzi'] = kanzi
            PostEntry['Contents'] = contents
            PostEntry['tags'] = nums
            PostEntry['date'] = datetime.today()

            # Post to MangoDB
            try:
                collection.insert_one(PostEntry)
                print('Posted: %s - %s' % (PostEntry['Kanzi'], PostEntry['Contents']))
            except pymongo.errors.DuplicateKeyError:
                #print('Failed to post %s , which is exist.' %PostEntry['Kanzi'])
                continue
    print('\n>>> All updates are uploaded to MongoDB. <<<')


# --- 3.8 Download Contents From MangoDB --- #
def MongoDB2Json_C(HostAddr, CollectionName):
    global ContentList

    # Connect to MongoDB
    client = pymongo.MongoClient(HostAddr)
    print('\nMaking a connection with MongoClient ...')
    db = client[DatabaseName]
    print('Getting Database - %s ...' % DatabaseName)
    collection = db[CollectionName]
    print('Getting a Collection - %s\n' % CollectionName)

    # Download all contents
    AllContents = collection.find()

    # Save to Json contents
    ContentList = {}
    for content in AllContents:
        EntryNum = content['tags']
        EntryKanzi = content['Kanzi']
        EntryContent = content['Contents']
        if EntryNum not in ContentList.keys():
            ContentList[EntryNum] = {}
        ContentList[EntryNum][EntryKanzi] = EntryContent

    # Save to Json file
    with open(FileJsonName_content, 'w') as cw:
        json.dump(ContentList, cw, indent=2)
    print('\n>>> All Data from MongoDB are saved to Json file. %s <<<' % FileJsonName_content)


# ###### Draw mails and send ###### #
# --- 4(1) Draw mails for each user --- #
def MailDraw():
    # Show List of User Info
    print('\n\n\n>>>>>> Mission Start - %s <<<<<<' % Today)
    print('\n=== Drawing mails for users as follows ===\n')
    IndexNo = 1
    for nums, details in UserList.items():
        print('No.%s : %s - %s ' % (str(IndexNo), nums, details['ID']))
        IndexNo += 1

    # Show List of Daily Contents
    #DrawCheckPoint = input('\nPlease check user list above, Enter Y to continue: ')
    DrawCheckPoint = 'Y'
    if DrawCheckPoint == 'Y':
        print('\n')
        for i, details in UserList.items():
            Receiver = Learner()
            StartDay = Receiver.GetStartDay(UserList, i)
            DayNumber = str(Receiver.GetDayNumber(StartDay, Today))
            TodayList = Receiver.GetTodayList(CheckList, DayNumber)
            MailTo, MailAddr = Receiver.GetMailInfo(UserList, i)
            print('%s : %s (%s) - %s' % (i, MailTo, MailAddr, TodayList))


# --- 4(2) Check and send mail to each user --- #
def MailSend():
    #SendCheckPoint = input('\nPlease check daily learning list of each user, Enter Y to send: ')
    SendCheckPoint = 'Y'
    if SendCheckPoint == 'Y':

        # Set mail box and log in
        from_addr = '37870979@qq.com'
        password = 'fcapyvxmjdbzbgji'
        smtp_server = 'smtp.qq.com'
        server = smtplib.SMTP_SSL(host=smtp_server)
        server.connect(smtp_server, 465)
        server.login(from_addr, password)

        # Set mail content
        for i in UserList.keys():
            print('\n\nCreating learner profile for %s...' % i)
            Receiver = Learner()

            print('Getting list of learning content ...')
            StartDay = Receiver.GetStartDay(UserList, i)
            DayNumber = str(Receiver.GetDayNumber(StartDay, Today))
            TodayList = Receiver.GetTodayList(CheckList, DayNumber)

            print('Getting lerning content of %s...' % TodayList)
            TodayNew, TodayReview = Receiver.GetTodayContent(ContentList, TodayList)

            print('Drawing E-mail ...')
            # Set mail text
            MailTo, MailAddr = Receiver.GetMailInfo(UserList, i)
            MailContent = Receiver.CreateMailContent(MailTo, TodayNew, TodayReview)
            msg = MIMEText(MailContent, 'plain', 'utf-8')

            # Set mail header
            msg['From'] = Header(from_addr)
            msg['To'] = MailAddr
            msg['Subject'] = Header('今日学习内容')

            # Send
            print('Mail to %s is sending ...' % i)
            server.sendmail(from_addr, MailAddr, msg.as_string())


        # Log out
        print('\n>>> All mails are sent! <<<')
        print(TodayDate)
        server.quit()

    else:
        pass


# ###### Load and save files ###### #
# --- Load Json file --- #
def LoadJsonFile():
    global CheckList, ContentList, UserList

    # Load Calendar from 'Calendar.json'
    try:
        with open(FileJsonName_date, 'r') as dr:
            CheckList = json.load(dr)
    except FileNotFoundError:
        pass

    # Load Contents from 'JLR_Contents.json'
    try:
        with open(FileJsonName_content, 'r') as cr:
            ContentList = json.load(cr)
    except FileNotFoundError:
        pass

    # Load User Info from 'UserInfo.json'
    try:
        with open(FileJsonName_user, 'r') as ur:
            UserList = json.load(ur)
    except FileNotFoundError:
        pass

    try:
        return CheckList, ContentList, UserList
    except NameError:
        pass


# --- Save as Json File --- #
def SaveJsonFile_C(ContentList):
    with open(FileJsonName_content, 'w') as cw:
        json.dump(ContentList, cw, indent=2)
    print('\n>>> All updates are saved. <<<')


# ###### Main Menu ###### #
def MainMenu():
    # Show Main Menu
    MenuRefresh = True
    while MenuRefresh:
        print('''
    ====== Welcome to Japanese Learning Reminder ======

            1. Save Calendar to Json File
            2. Edit Users
            3. Edit Contents

            4. Send Daily Learning Contents

            5. Quit

    ''')

        # Choose a function
        FunctionChoice = input('Please Enter the NUMBER of a function: ')

        # Save Calendar from Excel to Json
        if FunctionChoice == '1':
            CreateSchedule()

        # Edit Users
        if FunctionChoice == '2':
            print('''
    ====== Please Choose A Method of Users ======

        1. Show List of Users
        2. Add A New User

        3. Import User List from Excel to Json
        4. Export User List From Json to Excel

        5. Upload User List to MongoDB
        6. Download User List from MongoDB

        ''')
            MethodChoice_User = input('Please Enter the NUMBER of a function: ')

            # Show users list
            if MethodChoice_User == '1':
                ShowUserList()

            # Add a new user
            if MethodChoice_User == '2':
                AddNewUser()

            # User List from Excel to Json
            if MethodChoice_User == '3':
                Excel2Json_U()

            # User List from Json to Excel
            if MethodChoice_User == '4':
                Json2Excel_U()

            # Upload
            if MethodChoice_User == '5':
                Json2MongoDB_U(host_u, colname_u)

            # Download
            if MethodChoice_User == '6':
                MongoDB2Json_U(host_u, colname_u)

        # Edit Contents
        elif FunctionChoice == '3':
            print('''
    ====== Please Choose A Method of Contents ======

        1. Load and Inquire for Entries
        2. Input A New Entry
        3. Continue An Unfinished Entry
        4. Correct Values

        5. Import Contents From Excel to Json
        6. Export Contents From Json to Excel

        7. Upload Contents to MangoDB
        8. Download Contents From MangoDB

        ''')
            MethodChoice_Edit = input('Please Enter the NUMBER of a method: ')

            # Load and inquire Entries
            if MethodChoice_Edit == '1':
                Inquiry_C()

            # Input a New Entry
            elif MethodChoice_Edit == '2':
                AddNewEntry(ContentList)

            # Continue to input unfinished entry
            elif MethodChoice_Edit == '3':
                ContinueEntry(ContentList)

            # Correct Values
            elif MethodChoice_Edit == '4':
                SearchValue(ContentList)

            # Import contents from Excel
            elif MethodChoice_Edit == '5':
                Excel2Json_C()

            # Export to Excel
            elif MethodChoice_Edit == '6':
                Json2Excel_C(ContentList)

            # Upload to MangoDB
            elif MethodChoice_Edit == '7':
                Json2MongoDB_C(host_c, colname_c)

            # Download From MangoDB
            elif MethodChoice_Edit == '8':
                MongoDB2Json_C(host_c, colname_c)

        # Send Reminders
        elif FunctionChoice == '4':
            def AutoSend():
                MailDraw()
                MailSend()

            schedule.every().day.at("06:33").do(AutoSend)

            while True:
                schedule.run_pending()
                time.sleep(1)

        # Quit Main Menu
        elif FunctionChoice == '5':
            MenuRefresh = False


def Main():
    LoadJsonFile()
    MainMenu()


Main()
